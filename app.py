import os
import io
import time
import json
from datetime import datetime
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

import db_client
import scrapers
import query_strategist
import open_coding
import taxonomy_synthesizer
import classifier
import auditor
import orchestrator

load_dotenv()

# Set Streamlit Page Configuration
st.set_page_config(
    page_title="Blinkit Category-Discovery Engine",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom Styling (Charcoal dark mode with slate-green and soft yellow highlights matching Blinkit branding)
st.markdown("""
    <style>
    /* Hide sidebar nav */
    [data-testid="stSidebar"] {
        display: none !important;
    }
    
    /* Main App Container Styling */
    .stApp {
        background-color: #0b0f19 !important;
        color: #e2e8f0 !important;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    }
    
    /* Elegant Title Banner */
    .app-title {
        text-align: left;
        font-size: 2.3rem;
        font-weight: 700;
        letter-spacing: -0.5px;
        color: #ffcc00 !important;
        margin-top: 10px;
        margin-bottom: 2px;
    }
    .app-subtitle {
        text-align: left;
        font-size: 1.0rem;
        color: #00b560 !important;
        font-weight: 500;
        margin-bottom: 25px;
    }
    
    /* Metric Cards */
    [data-testid="stMetric"] {
        background-color: #111827 !important;
        border: 1px solid #1f2937 !important;
        border-radius: 8px !important;
        padding: 20px !important;
    }
    div[data-testid="stMetricValue"] {
        color: #f3f4f6 !important;
        font-weight: 700;
        font-size: 1.8rem !important;
    }
    div[data-testid="stMetricLabel"] {
        color: #9ca3af !important;
        font-weight: 500;
        font-size: 0.8rem !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    /* Buttons */
    .stButton>button {
        background-color: #1e293b !important;
        color: #ffffff !important;
        border: 1px solid #334155 !important;
        border-radius: 6px !important;
        font-weight: 500 !important;
        font-size: 0.85rem !important;
        padding: 8px 16px !important;
        transition: all 0.2s ease !important;
    }
    .stButton>button:hover {
        background-color: #00b560 !important;
        border-color: #00e676 !important;
        color: #ffffff !important;
    }
    
    /* Ingestion Alert dialog or container styling */
    div[data-testid="stNotification"] {
        background-color: #1e293b !important;
        border: 1px solid #334155 !important;
        border-radius: 8px !important;
        color: #ffffff !important;
    }
    
    /* Dataframes */
    div[data-testid="stDataFrame"] {
        border: 1px solid #1f2937;
        border-radius: 8px;
        background-color: #111827;
    }
    </style>
""", unsafe_allow_html=True)

# Helper to show active taxonomy categories in a modal popup
@st.dialog("📋 Active Feedback Categories & Definitions")
def show_taxonomy_popup():
    prop = taxonomy_synthesizer.load_taxonomy_proposal()
    if not prop or not prop.get("categories"):
        st.info("No active feedback categories discovered yet. Run classification first.")
    else:
        st.write("Below are the active feedback categories and definitions currently used by the classifier:")
        for c in prop.get("categories", []):
            st.markdown(f"**{c['name']}**")
            st.markdown(f"*{c['description']}*")
            if "examples" in c and c["examples"]:
                st.caption(f"Example Quote: \"{c['examples'][0]}\"")
            st.markdown("<div style='border-top:1px solid #1f2937; margin:10px 0;'></div>", unsafe_allow_html=True)

# Helper to load dataset
@st.cache_data(ttl=5)
def get_dashboard_data():
    return db_client.fetch_analyzed_data()

# Helper to generate Excel Sheet using openpyxl
def generate_excel_bytes(df):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Raw Ingestion & Analytics"
    ws2 = wb.create_sheet(title="PM Pivot Summary")
    
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=10)
    title_font = Font(name=font_family, size=14, bold=True, color="00b560")
    total_font = Font(name=font_family, size=10, bold=True)
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    accent_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Raw Data sheet headers and rows
    headers = list(df.columns)
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for index, row in df.iterrows():
        row_values = [str(val) if val is not None else "" for val in row]
        ws1.append(row_values)
        
    for r in range(2, ws1.max_row + 1):
        for c in range(1, ws1.max_column + 1):
            cell = ws1.cell(row=r, column=c)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if c != 4 else "left")
            
    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
        
    # Pivot sheet
    ws2.views.sheetView[0].showGridLines = True
    ws2["A2"] = "Blinkit Category-Discovery PM Pivot Report"
    ws2["A2"].font = title_font
    
    # Cross-tab table of Theme vs User Cohort
    if not df.empty:
        pivot = pd.crosstab(df["Theme"], df["User Type"], margins=True, margins_name="Total")
        
        ws2.cell(row=4, column=1, value="Theme vs Cohort Distribution").font = Font(name=font_family, size=12, bold=True)
        
        # Write headers
        ws2.cell(row=6, column=1, value="Theme / Cohort").font = header_font
        ws2.cell(row=6, column=1).fill = header_fill
        
        for col_idx, col_name in enumerate(pivot.columns, start=2):
            cell = ws2.cell(row=6, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Write rows
        current_row = 7
        for theme_name in pivot.index:
            cell = ws2.cell(row=current_row, column=1, value=theme_name)
            if theme_name == "Total":
                cell.font = total_font
                cell.fill = accent_fill
            else:
                cell.font = body_font
            cell.border = thin_border
            
            for col_idx, col_name in enumerate(pivot.columns, start=2):
                val = int(pivot.loc[theme_name, col_name])
                val_cell = ws2.cell(row=current_row, column=col_idx, value=val)
                val_cell.border = thin_border
                val_cell.alignment = Alignment(horizontal="right")
                if theme_name == "Total" or col_name == "Total":
                    val_cell.font = total_font
                    val_cell.fill = accent_fill
                else:
                    val_cell.font = body_font
            current_row += 1
            
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 5, 18)
        
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Dashboard Title
st.markdown("<h1 class='app-title'>Blinkit Category-Discovery Intelligence Engine</h1>", unsafe_allow_html=True)
st.markdown("<p class='app-subtitle'>Multi-Agent Analytical Pipeline diagnosing quick-commerce browse & category explore issues</p>", unsafe_allow_html=True)

# Fetch latest database records
df = get_dashboard_data()
db_counts = db_client.get_db_counts()

# ----------------------------------------------------
# Multi-Agent Phase Controller Panel
# ----------------------------------------------------
st.subheader("⚡ Analytics Pipeline Control Console")

# Database status metrics
col_m1, col_m2, col_m3 = st.columns(3)
with col_m1:
    m1_placeholder = st.empty()
    m1_placeholder.metric("📥 Unclassified Reviews in DB", db_counts["unclassified"])
with col_m2:
    m2_placeholder = st.empty()
    m2_placeholder.metric("🚀 Classified Reviews in DB", db_counts["classified"])
with col_m3:
    if st.button("🔄 Reset Feedback Categories", use_container_width=True):
        import os
        if os.path.exists("taxonomy_proposal.json"):
            try:
                os.remove("taxonomy_proposal.json")
            except Exception:
                pass
        st.success("Taxonomy reset!")
        time.sleep(1.0)
        st.rerun()
        
    if st.button("🗑️ Clear Classifications", use_container_width=True):
        db_client.clear_all_classifications()
        st.success("All classifications deleted!")
        time.sleep(1.0)
        st.rerun()
        
    if st.button("📋 View Feedback Categories", use_container_width=True):
        show_taxonomy_popup()

st.markdown("<br>", unsafe_allow_html=True)

col_panel1, col_panel2 = st.columns(2)

with col_panel1:
    with st.container(border=True):
        st.write("### 📥 Step 1: Scrape & Ingest Reviews")
        st.write("Extract customer feedback from the live App Store and Play Store.")
        scrape_limit = st.slider("Reviews to scrape", min_value=5, max_value=300, value=50, step=5)
        
        if st.button("📥 Start Scraping Run", use_container_width=True):
            progress_bar = st.progress(0.0)
            status_text = st.empty()
            
            status_text.markdown(f"🔄 **[1/2] Scraping stores** (limit: {scrape_limit})...")
            progress_bar.progress(0.4)
            play_records = scrapers.scrape_play_store(limit=scrape_limit)
            app_records = scrapers.scrape_app_store(limit=50)
            reddit_records = scrapers.scrape_reddit()
            all_ingested = play_records + app_records + reddit_records
            
            is_valid, msg = orchestrator.validate_ingestion(all_ingested)
            if is_valid:
                status_text.markdown("🔄 **[2/2] Updating filter keywords...**")
                progress_bar.progress(0.8)
                db_client.insert_raw_feedback(all_ingested)
                
                # Query Strategist
                db_client.log_pipeline_run("Targeted Query Selection", "STARTED")
                keywords = query_strategist.run_query_strategist(all_ingested[:100])
                db_client.log_pipeline_run("Targeted Query Selection", "COMPLETED", len(keywords))
                
                progress_bar.progress(1.0)
                status_text.markdown("✅ **Ingestion complete!**")
                st.success(f"Ingested {len(all_ingested)} real reviews!")
                time.sleep(1.5)
                st.rerun()
            else:
                db_client.log_pipeline_run("Feedback Ingestion", "FAILED", metadata={"error": msg})
                st.error(f"Scraping failed: {msg}")

with col_panel2:
    with st.container(border=True):
        st.write("### 🚀 Step 2: AI Classification & Auditing")
        st.write("Run the multi-agent classification pipeline on the unprocessed queue.")
        
        max_unclassified = db_counts["unclassified"]
        if max_unclassified == 0:
            st.info("📥 No reviews available to classify. Run Step 1 Scraping first.")
            classify_limit = 0
        else:
            default_val = min(20, max_unclassified)
            classify_limit = st.slider(
                "Reviews to classify", 
                min_value=1, 
                max_value=max_unclassified, 
                value=default_val, 
                step=1
            )
            # 3.0s throttle sleep per batch of 10 + ~0.5s network latency
            est_seconds = max(1.0, (classify_limit / 10.0) * 3.5)
            if est_seconds >= 60:
                time_str = f"{int(est_seconds // 60)}m {int(est_seconds % 60)}s"
            else:
                time_str = f"{int(est_seconds)}s"
            st.caption(f"⏱️ **Estimated Time to Finish**: {time_str} (10x faster batch mode!)")
            
        if classify_limit > 0 and st.button("🚀 Start AI Classification", use_container_width=True):
            progress_bar = st.progress(0.0)
            status_text = st.empty()
            
            # Auto-generate taxonomy if not exists
            prop = taxonomy_synthesizer.load_taxonomy_proposal()
            if not prop or not prop.get("categories"):
                status_text.markdown("🔄 **[1/3] Synthesizing taxonomy from raw reviews...**")
                progress_bar.progress(0.2)
                
                raw_reviews = db_client.fetch_unprocessed_feedback(limit=100)
                if not raw_reviews:
                    st.warning("Please ingest raw reviews first!")
                    st.stop()
                    
                themes = open_coding.run_open_coding(raw_reviews)
                prop = taxonomy_synthesizer.run_taxonomy_synthesis(themes, raw_reviews)
                prop["approved"] = True
                taxonomy_synthesizer.save_taxonomy_proposal(prop)
                
            categories = [c["name"] for c in prop["categories"]]
            
            status_text.markdown("🔄 **[2/3] Fetching unprocessed reviews...**")
            progress_bar.progress(0.4)
            unprocessed_records = db_client.fetch_unprocessed_feedback(limit=classify_limit)
            
            if not unprocessed_records:
                status_text.markdown("✅ **No unprocessed reviews to classify.**")
                progress_bar.progress(1.0)
                st.info("Database queue has 0 unclassified records.")
            else:
                db_client.log_pipeline_run("AI Classification", "STARTED")
                classified = []
                total_unprocessed = len(unprocessed_records)
                
                # Split into chunks of 10 for batch classification
                chunk_size = 10
                chunks = [unprocessed_records[k:k+chunk_size] for k in range(0, total_unprocessed, chunk_size)]
                total_chunks = len(chunks)
                
                for idx, chunk in enumerate(chunks):
                    # Progress bar scales from 0% to 90% during classification loop
                    pct = (float(idx+1) / total_chunks) * 0.9
                    progress_bar.progress(pct)
                    current_processed_count = min((idx + 1) * chunk_size, total_unprocessed)
                    remaining_reviews = total_unprocessed - current_processed_count
                    status_text.markdown(f"⚡ **Classifying reviews [{current_processed_count}/{total_unprocessed}]** ({remaining_reviews} remaining)...")
                    
                    batch_res = classifier.classify_reviews_batch(chunk, categories)
                    
                    # Failsafe chunk saving: validate and insert each batch of 10
                    is_valid, msg = orchestrator.validate_classification(batch_res, categories)
                    if is_valid:
                        db_client.insert_ai_analytics(batch_res)
                        classified.extend(batch_res)
                        
                        # Dynamically update the metrics on the screen!
                        cur_class = db_counts["classified"] + len(classified)
                        cur_unclass = max(0, db_counts["unclassified"] - len(classified))
                        m1_placeholder.metric("📥 Unclassified Reviews in DB", cur_unclass)
                        m2_placeholder.metric("🚀 Classified Reviews in DB", cur_class)
                    else:
                        st.warning(f"Batch validation warning in batch {idx+1}: {msg}. Retrying items individually...")
                        valid_singles = []
                        for single in batch_res:
                            is_single_valid, _ = orchestrator.validate_classification([single], categories)
                            if is_single_valid:
                                valid_singles.append(single)
                        if valid_singles:
                            db_client.insert_ai_analytics(valid_singles)
                            classified.extend(valid_singles)
                            
                            # Dynamically update the metrics on the screen!
                            cur_class = db_counts["classified"] + len(classified)
                            cur_unclass = max(0, db_counts["unclassified"] - len(classified))
                            m1_placeholder.metric("📥 Unclassified Reviews in DB", cur_unclass)
                            m2_placeholder.metric("🚀 Classified Reviews in DB", cur_class)
                    
                    # Throttle between batch API calls
                    if classifier.GROQ_API_KEY and idx < total_chunks - 1:
                        time.sleep(3.0)
                
                if classified:
                    # Consensus Auditing
                    status_text.markdown("🔄 **[3/3] Running Consensus Auditing verification...**")
                    progress_bar.progress(0.9)
                    db_client.log_pipeline_run("AI Classification", "COMPLETED", len(classified))
                    
                    rate, audited = auditor.run_auditor(classified, categories)
                    db_client.log_pipeline_run("Consensus Auditing", "COMPLETED", len(audited), {"agreement_rate": rate})
                    
                    progress_bar.progress(1.0)
                    status_text.markdown(f"✅ **Classification & Auditing complete!** Agreement: {rate:.1%}")
                    st.success(f"Classified {len(classified)} reviews! Auditor agreement: {rate:.1%}")
                    time.sleep(2.0)
                    st.rerun()
                else:
                    db_client.log_pipeline_run("AI Classification", "FAILED", metadata={"error": "Zero reviews successfully validated."})
                    st.error("Classification failed: No records successfully validated and saved.")

# ----------------------------------------------------
# KPIs & Analytical Metrics
# ----------------------------------------------------
st.markdown("<hr style='border:0; border-top:1px solid #1f2937; margin:20px 0;'>", unsafe_allow_html=True)
st.subheader("📊 Key Exploration Insights & Metrics")

if not df.empty:
    total_reviews = len(df)
    frustrated_pct = len(df[df["Sentiment"].isin(["Highly Frustrated", "Disappointed"])]) / total_reviews if total_reviews > 0 else 0
    top_theme = df["Theme"].value_counts().index[0] if total_reviews > 0 else "N/A"
    
    # Inter-agent agreement rate from pipeline run logs
    runs = db_client.fetch_pipeline_runs(10)
    auditor_runs = [r for r in runs if r["phase"] == "Consensus Auditing" and r["status"] == "COMPLETED"]
    if auditor_runs:
        last_val = json.loads(auditor_runs[0]["validation_results"]) if isinstance(auditor_runs[0]["validation_results"], str) else auditor_runs[0]["validation_results"]
        agreement_rate = f"{last_val.get('agreement_rate', 0.0):.1%}" if last_val else "N/A"
    else:
        agreement_rate = "90.0% (Simulated)"
        
    # Spot-check validity calculation
    spot_checked_df = df[df["Spot Checked"] == True]
    valid_spot_checked = spot_checked_df[spot_checked_df["Spot Check Valid"] == True]
    if len(spot_checked_df) > 0:
        spot_check_rate = f"{len(valid_spot_checked) / len(spot_checked_df):.1%}"
    else:
        spot_check_rate = "N/A (Awaiting Check)"
else:
    total_reviews = 0
    frustrated_pct = 0
    top_theme = "N/A"
    agreement_rate = "N/A"
    spot_check_rate = "N/A"

kpi_c1, kpi_c2, kpi_c3, kpi_c4 = st.columns(4)
with kpi_c1:
    st.metric("Reviews Classified", total_reviews)
with kpi_c2:
    st.metric("User Frustration Rate", f"{frustrated_pct:.1%}")
with kpi_c3:
    st.metric("Primary explore pain point", top_theme)
with kpi_c4:
    st.metric("Auditor Agreement Rate", agreement_rate)

# ----------------------------------------------------
# Visualizations
# ----------------------------------------------------
if not df.empty:
    col_chart1, col_chart2 = st.columns(2)
    with col_chart1:
        st.markdown("**Category explore Pain Points Distribution**")
        theme_counts = df["Theme"].value_counts().reset_index()
        theme_counts.columns = ["Theme", "Count"]
        st.bar_chart(theme_counts.set_index("Theme"))
    with col_chart2:
        st.markdown("**Affected User Cohort Segments**")
        cohort_counts = df["User Type"].value_counts().reset_index()
        cohort_counts.columns = ["Cohort", "Count"]
        st.bar_chart(cohort_counts.set_index("Cohort"))

# ----------------------------------------------------
# Human Checkpoint: Manual Spot Check Console
# ----------------------------------------------------
st.markdown("<hr style='border:0; border-top:1px solid #1f2937; margin:20px 0;'>", unsafe_allow_html=True)
st.subheader("🎯 Human Spot-Checking & Validation Console")

# Fetch records flagged for spot check where validation has not been decided yet
if not df.empty and "Spot Checked" in df.columns:
    spot_check_queue = df[(df["Spot Checked"] == True) & (df["Spot Check Valid"].isna())]

    if spot_check_queue.empty:
        st.info("🎉 All spot checks are complete. No pending records to verify.")
    else:
        st.write(f"The Auditor flagged {len(spot_check_queue)} records for validation. Please confirm if the classification is correct:")
        
        current_check = spot_check_queue.iloc[0]
        review_id = current_check["Review ID"]
        
        with st.container(border=True):
            st.write(f"**Review Text:** \"{current_check['Text']}\"")
            st.write(f"**Source:** {current_check['Source']} | **App Version:** {current_check['App Version']}")
            if current_check['Theme'] == "Ineligible / AI Failure":
                st.markdown(f"**AI Theme Classification:** :red[{current_check['Theme']}] ⚠️ *AI classification failed.*")
            else:
                st.write(f"**AI Theme Classification:** `{current_check['Theme']}`")
            st.write(f"**AI Sentiment Classification:** `{current_check['Sentiment']}`")
            st.write(f"**AI Cohort Classification:** `{current_check['User Type']}`")
            
            # Load active categories list for manual corrections
            categories_list = []
            prop = taxonomy_synthesizer.load_taxonomy_proposal()
            if prop and prop.get("categories"):
                categories_list = [c["name"] for c in prop["categories"]]
            if not categories_list:
                categories_list = ["Pricing & Refund Issues", "Product Quality & Freshness", "Delivery Speed & Delay", "App Navigation & Clutter", "Customer Support Issues"]
            
            st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
            correct_cat = st.selectbox(
                "💡 If invalid, select the correct category to reassign:",
                options=categories_list,
                index=categories_list.index(current_check["Theme"]) if current_check["Theme"] in categories_list else 0
            )
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("✅ Confirm Classification as Valid", use_container_width=True):
                    db_client.update_spot_check(review_id, True)
                    st.success("Validated classification!")
                    st.cache_data.clear()
                    st.rerun()
            with col_btn2:
                if st.button("❌ Flag & Reassign Category", use_container_width=True):
                    db_client.update_classification_category(review_id, correct_cat)
                    st.warning(f"Flagged as invalid and reassigned to '{correct_cat}'!")
                    st.cache_data.clear()
                    st.rerun()
else:
    st.info("📋 No classified reviews in database to spot check. Run the classification pipeline first.")

# Display spot check stats
if total_reviews > 0:
    st.write(f"**Human Spot-Check Agreement Rate:** {spot_check_rate} (Valid spot-checks out of completed checks)")

# ----------------------------------------------------
# Export & Reporting Container
# ----------------------------------------------------
st.markdown("<hr style='border:0; border-top:1px solid #1f2937; margin:20px 0;'>", unsafe_allow_html=True)
st.subheader("📥 Export Reports")

with st.container(border=True):
    col_d1, col_d2 = st.columns([3, 1])
    with col_d1:
        st.markdown("<p style='margin:0; font-size:0.9rem; color:#9ca3af;'>Generate a comprehensive Microsoft Excel report. Includes all raw feedback, metadata columns, and a cross-tabulated <b>PM Pivot Summary</b> sheet comparing Dynamic Themes vs User Cohort segments.</p>", unsafe_allow_html=True)
    with col_d2:
        if not df.empty:
            excel_bin = generate_excel_bytes(df)
            st.download_button(
                label="💾 DOWNLOAD EXCEL REPORT",
                data=excel_bin,
                file_name=f"Blinkit_Category_Discovery_Metrics_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.button("💾 DOWNLOAD EXCEL REPORT", disabled=True, use_container_width=True)

# ----------------------------------------------------
# Classified Feed
# ----------------------------------------------------
st.markdown("<hr style='border:0; border-top:1px solid #1f2937; margin:20px 0;'>", unsafe_allow_html=True)
st.subheader("📋 Classified Reviews Feed")
if not df.empty:
    st.dataframe(
        df[["Timestamp", "Source", "App Version", "Theme", "Sentiment", "User Type", "Root Cause", "Text"]],
        use_container_width=True,
        column_config={
            "Timestamp": st.column_config.DatetimeColumn(format="YYYY-MM-DD HH:mm"),
            "Text": st.column_config.TextColumn(width="large")
        }
    )
else:
    st.info("No records classified yet. Run the pipeline control panel above to populate analytics.")
